#https://stackoverflow.com/questions/51800122/using-openpyxl-to-find-rows-that-contain-cell-with-specific-value-python-3-6/51801421
# component_code_list = [STT in Excel 1, Component Code, STT in Excel 2, Ma vat tu cu]
# first excel sheet name
# wb range for first sheet
from openpyxl import Workbook
import openpyxl
print("SAP B1 auto tool for TSAN")
# Specify the first file, where we will add a blank column and fill values
excel_path_suffix = "./../thong_dmvt_lo1048/file_goc/"
file = excel_path_suffix + "PL02, PL03_VTLK (Xong theo Ebom 11 du kien)_main.xlsx"
# Select a specific of the first excel file's sheet to work
new_bom_sheet = "VTLK Thau"
# (we don't handle the first info rows in the excel sheet because it's kind of time-consuming for now)
# specify item range where the real items are available
wb_first_row = 3
wb_last_row = 169

print("Opening excel file: " + str(file))
new_bom_wb = openpyxl.load_workbook(file)

ws = new_bom_wb[new_bom_sheet]
print("Working with excel sheet: " + str(ws))

component_code_list = []
# iterate every row in column variable
i = 1
# limit to the rows where Component code are available 

for row in ws.iter_rows("H"):
    for cell in row:
        # note that some cases we must manually link in order to complete an Excel sheet. 
        if i>=wb_first_row and i<=wb_last_row:
            component_code_list.append([i,cell.value])
        i = i+1
print("There're total of " + str(wb_last_row-wb_first_row+1) + " component to be map the new SAP B1 component code")
# print the list contain all ma nxs to map to target file along with its row index in previous file to Danh muc vat tu LINH KIEN.xlsx
# specify the second excel file
file1 = excel_path_suffix + "Danh muc vat tu LINH KIEN.xlsx"
print("Opening excel file: " + str(file1))
wb1 = openpyxl.load_workbook(file1, read_only=True)
ws1 = wb1.active
print("Working with excel sheet: " + str(wb1.sheetnames[0]))
component_code_scanned_row_index = 1
# search each item in component_code_list in every row of column C (Ky ma hieu) in `Danh muc vat tu LINH KIEN.xlsx`
# counter the number of original component to be searched
component_code_cnt = 0
none1_cnt = 0
list_has_foreign_name = []
# detect how many matches are there for 1 input after scanning all the row on target column
componentcode_match_in_loop_cnt = 0
match_indices = []
for item in component_code_list:
    print("------")
    # reset the counter for next row scanning loop    
    component_code_scanned_row_index = 1
    for row in ws1.iter_rows("C"):
        for cell in row:
            # if we detect any matching by component code
            if cell.value == item[1]:
                componentcode_match_in_loop_cnt = componentcode_match_in_loop_cnt + 1
                # print("STT in second file: "+ str(component_code_scanned_row_index))
                match_indices.append(component_code_scanned_row_index)
               # print(component_code_scanned_row_index)
               # print("Foreign Name (Ma cu): " + str(ws1.cell(None, component_code_scanned_row_index, 4).value))
                # only consider the first match
                if componentcode_match_in_loop_cnt == 1:
                    component_code_cnt = component_code_cnt+1
                    item.append(component_code_scanned_row_index)
                    item.append(ws1.cell(None, component_code_scanned_row_index, 4).value)
                    if ws1.cell(None, component_code_scanned_row_index, 4).value == None:
                        print("STT in first file: " + str(item[0]))
                        print("None value detected, we won't use this for the next code")
                        print("The component in old BOM match current new BOM, but its Foreign Name is None")
                        none1_cnt = none1_cnt+1
                    else:
                        list_has_foreign_name.append(item)
            component_code_scanned_row_index = component_code_scanned_row_index+1
    print("STT in first file: " + str(item[0]))
    print("Component code (Ma NXS1): " + str(item[1]))

    if componentcode_match_in_loop_cnt != 0:
        print("Detect " + str(componentcode_match_in_loop_cnt) + " matching in second file With these row corresponding indices: ")
        print(match_indices)
        match_indices = []
    else:
        print("Didn't detect any matching")
        #print("It means that it doesn't have new sap b1 component code")
    componentcode_match_in_loop_cnt = 0

print("------")
print("Number of items matched by component code and have Foreign Name: " + str(component_code_cnt-none1_cnt) + " / " + str(wb_last_row-wb_first_row+1))
nsx_code_match_row_index = 1
file2 = excel_path_suffix + "List of Items 31-03-2020.xlsx"
print("Opening excel file: " + str(file2))
wb2 = openpyxl.load_workbook(file2, read_only=True)
ws2 = wb2.active
print("Working with excel sheet: "+str(wb2.sheetnames[0]))
match_by_oldsapcode_cnt = 0
old_sap_code_match_in_loop_cnt = 0
process_item_cnt = 0
match_by_oldsapcode_indices = []
for item in list_has_foreign_name:
    print("------")
    process_item_cnt = process_item_cnt + 1
    for row in ws2.iter_rows("F"):
        for cell in row:
            if cell.value == item[3]:
                old_sap_code_match_in_loop_cnt = old_sap_code_match_in_loop_cnt + 1
                match_by_oldsapcode_indices.append(nsx_code_match_row_index)
                # only consider the first match
                if old_sap_code_match_in_loop_cnt == 1:
                    match_by_oldsapcode_cnt = match_by_oldsapcode_cnt + 1
                    if ws2.cell(None, nsx_code_match_row_index, 2).value == None:
                        print("None value detected")
                    # STT in third excel
                    item.append(nsx_code_match_row_index)
                    item.append(ws2.cell(None, nsx_code_match_row_index, 2).value)
                    # Write Ma cu
                    ws.cell(row=item[0], column=4).value = item[3]
                    # Ma moi
                    ws.cell(row=item[0], column=5).value = ws2.cell(None, nsx_code_match_row_index, 2).value
            nsx_code_match_row_index = nsx_code_match_row_index+1

    print("Processed item: " + str(process_item_cnt) + " / " + str(len(list_has_foreign_name)))
    print("STT in first file: " + str(item[0]))
    print("Component code (Ma NXS1): " + str(item[1]))
    print("STT in second file: " + str(item[2]))
    print("Old SAP B1 code: " + str(item[3]))
    if old_sap_code_match_in_loop_cnt != 0:
        print("Detect " + str(old_sap_code_match_in_loop_cnt) + " matching in second file With these row with following indices: ")
        print(match_by_oldsapcode_indices)
        match_by_oldsapcode_indices = []
    else:
        print("Didn't detect old_sap_code_match_in_loop_cnt matching")
        print("It means that it doesn't have new sap b1 component code")
        ws.cell(row=item[0], column=5).value = "ma_moi"
    old_sap_code_match_in_loop_cnt = 0
    # reset the row counter for next matching loop    
    nsx_code_match_row_index = 1
print("There are: " + str(len(list_has_foreign_name)-match_by_oldsapcode_cnt) + " component without new sap b1 code")
#print(component_code_list)
des_file = excel_path_suffix+'sample_book_new_file_22.xlsx'
print("Saving the file to location: " + str(des_file))
new_bom_wb.save(filename = des_file)
print("Finish excel mapping automation tool")