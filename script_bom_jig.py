#!/usr/bin/python
# Find the matched column by the componnent code, if the partnumber in the first excel file is a substring of the string of items in List (second excel)
from openpyxl import Workbook
import openpyxl

DEFINE_SHEET = 2
# Python program to print 
# colored text and background 
def prRed(skk): print("\033[91m {}\033[00m" .format(skk)) 
def prGreen(skk): print("\033[92m {}\033[00m" .format(skk)) 
def prYellow(skk): print("\033[93m {}\033[00m" .format(skk)) 
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk)) 
def prPurple(skk): print("\033[95m {}\033[00m" .format(skk)) 
def prCyan(skk): print("\033[96m {}\033[00m" .format(skk)) 
def prLightGray(skk): print("\033[97m {}\033[00m" .format(skk)) 
def prBlack(skk): print("\033[98m {}\033[00m" .format(skk)) 

prRed("Check output SAP B1 auto tool for TSAN")

# FIRST EXCEL FILE
excel_path_suffix = "./../tailieu_machjig/"
if DEFINE_SHEET == 1:
    file = excel_path_suffix + "general_jig/General_Jig_V3.xlsx"
elif DEFINE_SHEET == 2:
    file = excel_path_suffix + "videoengine_jig/VIDEOENGINE_JIG_V3.xlsx"

new_bom_wb = openpyxl.load_workbook(file)

# USER DEFINES

# Select the linking column between file 1 and file 2
if DEFINE_SHEET == 1:
    new_bom_sheet = "General_Jig_V2"
    file0_linking_col = "E"
elif DEFINE_SHEET == 2:
    # Select a specific sheet to work
    new_bom_sheet = "VIDEOENGINE_JIG_V2"
    file0_linking_col = "C"

# limit to the rows where Component code are available 
# (we don't handle the first info rows in the excel sheet because it's kind of time-consuming for now)
# in future we could use automatic function in order to detect range where they're actual items
if DEFINE_SHEET == 1:
    wb_first_row = 3
    wb_last_row = 97
elif DEFINE_SHEET == 2:
    wb_first_row = 2
    wb_last_row = 40
# Print information
ws = new_bom_wb[new_bom_sheet]
prGreen("Opening excel file: " + str(file))
prGreen("Working with excel sheet: " + str(ws))
prGreen("Column in file 1 where link file 1 and 2: " + file0_linking_col)
# PROGRAMMING VARIABLES
# iterate every row in column variable
i = 1
component_code_list = []
for row in ws.iter_rows(file0_linking_col):
    for cell in row:
        # note that some cases we must manually link in order to complete an Excel sheet. 
        if i >= wb_first_row and i <= wb_last_row:
            component_code_list.append([i,cell.value])
        i = i+1
print(component_code_list)
# SECOND EXCEL FILE
file2 = excel_path_suffix + "List of Items_4_6_2020.xlsx"

wb2 = openpyxl.load_workbook(file2, read_only=True)
ws2 = wb2.active

# USER DEFINES
two_and_one_col = "C"
# print
prGreen("Opening excel file: " + str(file2))
prGreen("Working with excel sheet: "+str(wb2.sheetnames[0]))
prGreen("Column in file 2 where link file 1 and 2: " + two_and_one_col)
# PROGRAMMING VARIABLES
nsx_code_match_row_index = 1
match_by_oldsapcode_cnt = 0
old_sap_code_match_in_loop_cnt = 0
process_item_cnt = 0
match_by_partnumber_indices = []
list_has_foreign_name = []
def print_info_3():
    print("Processed item: " + str(process_item_cnt) + " / " + str(len(component_code_list)))
    print("STT in first file: " + str(item[0]))
    print("Part Number: " + str(item[1]))
# write to first original file, at the matched row
# in "specific chosen colum" with the value of input chosen column in second file 
def write_to_excel(row_in_first_file, row_in_second_file):
    # O clumn
    first_column = 0
    second_column = 0
    if DEFINE_SHEET == 1:
        first_column = 8
        second_column = 10
    elif DEFINE_SHEET == 2:
        first_column = 8
        second_column = 18
    # Write SAP B1 code in ERP system
    ws.cell(row=row_in_first_file, column=first_column).value = ws2.cell(None, row_in_second_file, 2).value

number_of_match_morethanone = 0
for item in component_code_list:
    prRed("----------------------")
    process_item_cnt = process_item_cnt + 1
    # Write new sap code column, the column links one and second files
    # ws.cell(row=item[0], column=new_code_col).value = item[1]
    for row in ws2.iter_rows(two_and_one_col):
        for cell in row:
            # continue to next loop if variable is empty
            if item[1] == "":
                continue
            if (cell.value.find(item[1]) != -1):
                old_sap_code_match_in_loop_cnt = old_sap_code_match_in_loop_cnt + 1
                match_by_partnumber_indices.append(nsx_code_match_row_index)
                # only consider the first match
                if old_sap_code_match_in_loop_cnt == 1:
                    match_by_oldsapcode_cnt = match_by_oldsapcode_cnt + 1
                    # if ws2.cell(None, nsx_code_match_row_index, 2).value == None:
                    #     print("None value detected")
                    # STT in second excel, where the match happens
                    item.append(nsx_code_match_row_index)
                    item.append(ws2.cell(None, nsx_code_match_row_index, 6).value)
                    # Write to excel
                    write_to_excel(item[0], nsx_code_match_row_index)
            nsx_code_match_row_index = nsx_code_match_row_index+1

    print_info_3()
    if old_sap_code_match_in_loop_cnt != 0:
        print("Detect " + str(old_sap_code_match_in_loop_cnt) + " matching in second file With these row with following indices: ")
        print(match_by_partnumber_indices)
        if len(match_by_partnumber_indices)>1:
            number_of_match_morethanone = number_of_match_morethanone+1
        match_by_partnumber_indices = []
    else:
        print("Didn't detect old_sap_code_match_in_loop_cnt matching")
        print("It means that it doesn't have new sap b1 component code")
        #ws.cell(row=item[0], column=new_code_col).value = "no_info"
    old_sap_code_match_in_loop_cnt = 0
    # reset the row counter for next matching loop    
    nsx_code_match_row_index = 1
print("There are: " + str(len(component_code_list)-match_by_oldsapcode_cnt) + " component aren't in the ERP system")
print(component_code_list)

if DEFINE_SHEET == 1:
    output_unique_str = "general_jig"
if DEFINE_SHEET == 2:
    output_unique_str = "video_jig"

des_file = excel_path_suffix + output_unique_str + '_out' + '.xlsx'
print("Saving the file to location: " + str(des_file))
new_bom_wb.save(filename = des_file)
print("Finish excel mapping automation tool")
print(number_of_match_morethanone)
