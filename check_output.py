#!/usr/bin/python
# This program double checks the output Excel file and list of items in the database. Make sure that no mistakes were made

# https://stackoverflow.com/questions/51800122/using-openpyxl-to-find-rows-that-contain-cell-with-specific-value-python-3-6/51801421
# https://www.geeksforgeeks.org/print-colors-python-terminal/
# component_code_list = [STT in Excel 1, Component Code, STT in Excel 2, Ma vat tu cu]
# first excel sheet name
# wb range for first sheet
#
from openpyxl import Workbook
import openpyxl

# Python program to print 
# colored text and background 
def prRed(skk): print("\033[91m {}\033[00m" .format(skk)) 
def prGreen(skk): print("\033[92m {}\033[00m" .format(skk)) 
def prYellow(skk): print("\033[93m {}\033[00m" .format(skk)) 
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk)) 
def prPurple(skk): print("\033[95m {}\033[00m" .format(skk)) 
def prCyan(skk): print("\033[96m {}\033[00m" .format(skk)) 
def prLightGray(skk): print("\033[97m {}\033[00m" .format(skk)) 
def prBlack(skk): print("\033[98m {}\033[00m" .format(skk)) 

prRed("Check output SAP B1 auto tool for TSAN")

# FIRST EXCEL FILE
excel_path_suffix = "./../thong_dmvt_lo1048/file_goc/"

file = excel_path_suffix + "PL02, PL03_VTLK (Xong theo Ebom 11 du kien)_v3.1_script.xlsx"
new_bom_wb = openpyxl.load_workbook(file)

# USER DEFINES
# Select a specific sheet to work
new_bom_sheet = "Vat tu chi dinh"
# Select the linking column between file 1 and file 2
file0_linking_col = "F"
# limit to the rows where Component code are available 
# (we don't handle the first info rows in the excel sheet because it's kind of time-consuming for now)
# in future we could use automatic function in order to detect range where they're actual items
wb_first_row = 7
wb_last_row = 31

# Print information
ws = new_bom_wb[new_bom_sheet]
prGreen("Opening excel file: " + str(file))
prGreen("Working with excel sheet: " + str(ws))
prGreen("Column in file 1 where link file 1 and 2: " + file0_linking_col)
# PROGRAMMING VARIABLES
# iterate every row in column variable
i = 1
component_code_list = []
for row in ws.iter_rows(file0_linking_col):
    for cell in row:
        # note that some cases we must manually link in order to complete an Excel sheet. 
        if i >= wb_first_row and i <= wb_last_row:
            component_code_list.append([i,cell.value.strip()])
        i = i+1

# SECOND EXCEL FILE
file2 = excel_path_suffix + "List of Items_7.4.2020.xlsx"

wb2 = openpyxl.load_workbook(file2, read_only=True)
ws2 = wb2.active

# USER DEFINES
two_and_one_col = "B"
old_code_col = 13
new_code_col = 7
# print
prGreen("Opening excel file: " + str(file2))
prGreen("Working with excel sheet: "+str(wb2.sheetnames[0]))
prGreen("Column in file 2 where link file 1 and 2: " + two_and_one_col)
# PROGRAMMING VARIABLES
nsx_code_match_row_index = 1
match_by_oldsapcode_cnt = 0
old_sap_code_match_in_loop_cnt = 0
process_item_cnt = 0
match_by_oldsapcode_indices = []
list_has_foreign_name = []
def print_info_3():
    print("Processed item: " + str(process_item_cnt) + " / " + str(len(component_code_list)))
    print("STT in first file: " + str(item[0]))
    print("New ma VTTH to checked: " + str(item[1]))
    
for item in component_code_list:
    prRed("----------------------")
    process_item_cnt = process_item_cnt + 1
    # Write new sap code column, the column links one and second files
    # ws.cell(row=item[0], column=new_code_col).value = item[1]
    for row in ws2.iter_rows(two_and_one_col):
        for cell in row:
            if cell.value == item[1].strip():
                old_sap_code_match_in_loop_cnt = old_sap_code_match_in_loop_cnt + 1
                match_by_oldsapcode_indices.append(nsx_code_match_row_index)
                # only consider the first match
                if old_sap_code_match_in_loop_cnt == 1:
                    match_by_oldsapcode_cnt = match_by_oldsapcode_cnt + 1
                    # if ws2.cell(None, nsx_code_match_row_index, 2).value == None:
                    #     print("None value detected")
                    # STT in second excel, where the match happens
                    item.append(nsx_code_match_row_index)
                    item.append(ws2.cell(None, nsx_code_match_row_index, 6).value)

                    # Write ma cu lay tu List of Item to first excel file
                    ws.cell(row=item[0], column=15).value = ws2.cell(None, nsx_code_match_row_index, 6).value
                    # Description
                    ws.cell(row=item[0], column=16).value = ws2.cell(None, nsx_code_match_row_index, 3).value
            nsx_code_match_row_index = nsx_code_match_row_index+1

    print_info_3()
    if old_sap_code_match_in_loop_cnt != 0:
        print("Detect " + str(old_sap_code_match_in_loop_cnt) + " matching in second file With these row with following indices: ")
        print(match_by_oldsapcode_indices)
        match_by_oldsapcode_indices = []
    else:
        print("Didn't detect old_sap_code_match_in_loop_cnt matching")
        print("It means that it doesn't have new sap b1 component code")
        ws.cell(row=item[0], column=new_code_col).value = "ma_moi"
    old_sap_code_match_in_loop_cnt = 0
    # reset the row counter for next matching loop    
    nsx_code_match_row_index = 1
print("There are: " + str(len(component_code_list)-match_by_oldsapcode_cnt) + " component without new sap b1 code")
print(component_code_list)
des_file = excel_path_suffix +'VTCD2.xlsx'
print("Saving the file to location: " + str(des_file))
new_bom_wb.save(filename = des_file)
print("Finish excel mapping automation tool")

